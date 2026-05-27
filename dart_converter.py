"""
DART → K-IFRS 자동 컨버전 파이프라인
=====================================
기업명만 입력하면 DART에서 데이터를 수집하고 K-IFRS 전환 엑셀을 자동 생성한다.

흐름:
  기업명 입력
    → DART 검색 (corp_code 확인)
    → 기업개황 조회 (업종, 결산월)
    → 회계기준 확인 (K-GAAP인지 검증)
    → 재무제표 수치 수집 (BS/IS/CF 상세)
    → 감사보고서 검색
    → 전환근거 자동 매핑
    → 전환 변동내역 산출
    → 엑셀 출력
"""

import os
import re
import json
from datetime import datetime
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field

from dart_api import DartAPI, CompanyInfo, FinancialAccount, DisclosureInfo, DartAPIError


# ═══════════════════════════════════════════════════════════
# 전환 결과 데이터 구조
# ═══════════════════════════════════════════════════════════

@dataclass
class ConversionItem:
    """전환근거 항목"""
    no: int
    category: str                  # 자산/부채/자본/수익/비용
    account_name: str              # K-GAAP 계정과목
    book_value: Optional[int]      # 장부금액
    kifrs_standard: str            # 적용 K-IFRS (예: "1116")
    kifrs_name: str                # K-IFRS 기준서명
    change_description: str        # 전환 시 변동사항
    impact: str                    # HIGH/MEDIUM/LOW/NONE
    accuracy: str = "★★☆"        # 정확도 (DART 기반이므로 기본 ★★☆)
    note: str = ""


@dataclass
class ConversionDelta:
    """전환 변동내역 항목"""
    account_name: str
    kgaap_amount: Optional[int]
    adjustment: Optional[int]
    kifrs_estimated: Optional[int]
    changed: bool
    kifrs_standard: str
    change_basis: str              # 변동/무변동 근거
    assumptions: str = ""          # 추정 가정


@dataclass
class ConversionResult:
    """전환 결과 전체"""
    company: CompanyInfo
    fiscal_year: str
    accounting_standard: str       # K-GAAP / K-IFRS
    accuracy_grade: str            # D등급 (DART 공시 기반)
    bs_accounts: List[FinancialAccount] = field(default_factory=list)
    is_accounts: List[FinancialAccount] = field(default_factory=list)
    cf_accounts: List[FinancialAccount] = field(default_factory=list)
    all_accounts: List[FinancialAccount] = field(default_factory=list)
    conversion_items: List[ConversionItem] = field(default_factory=list)
    conversion_deltas: List[ConversionDelta] = field(default_factory=list)
    checklist: List[Dict] = field(default_factory=list)
    audit_report: Optional[DisclosureInfo] = None
    warnings: List[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════
# K-IFRS 매핑 규칙 (auto_mapping_logic.md 기반)
# ═══════════════════════════════════════════════════════════

# 계정과목명 → (K-IFRS 기준서번호, 기준서명, 기본 영향도, 기본 변동사항)
ACCOUNT_KIFRS_MAP = [
    # ── 자산 ──
    (r'현금|현금성자산', '1007', 'IAS 7 현금흐름표', 'LOW', '명칭 및 표시 변경 가능'),
    (r'단기금융상품|정기예금|정기적금', '1109', 'IFRS 9 금융상품', 'MEDIUM', 'AC/FVPL/FVOCI 재분류 필요'),
    (r'매출채권', '1109', 'IFRS 9 금융상품', 'HIGH', 'ECL(기대신용손실) 모형 적용 필수'),
    (r'대손충당금', '1109', 'IFRS 9 금융상품', 'HIGH', 'ECL 산출로 대체'),
    (r'미수금|미수수익', '1109', 'IFRS 9 금융상품', 'LOW', 'AC 측정, ECL 적용 여부 검토'),
    (r'선급금|선급비용', '1001', 'IAS 1 재무제표 표시', 'LOW', '표시 변경'),
    (r'재고자산|상품$|제품$|원재료', '1002', 'IAS 2 재고자산', 'MEDIUM', 'NRV 평가, 후입선출법 불허'),
    (r'장기금융상품', '1109', 'IFRS 9 금융상품', 'MEDIUM', 'AC/FVPL/FVOCI 재분류'),
    (r'매도가능|투자증권|지분증권', '1109', 'IFRS 9 금융상품', 'HIGH', 'FVOCI/FVPL 재분류, OCI 처리 변경'),
    (r'대여금', '1109', 'IFRS 9 금융상품', 'LOW', 'AC 측정, ECL 적용'),
    (r'토지$', '1016', 'IAS 16 유형자산', 'MEDIUM', '재평가모형 선택 가능, 손상차손 검토'),
    (r'건물$', '1016', 'IAS 16 유형자산', 'MEDIUM', '잔존가치·내용연수 재검토, 구성요소별 감가상각'),
    (r'기계장치|차량|비품|공구', '1016', 'IAS 16 유형자산', 'LOW', '잔존가치·내용연수 재검토'),
    (r'건설중인자산', '1016', 'IAS 16 유형자산', 'LOW', '차입원가 자본화(IAS 23) 검토'),
    (r'무형자산|영업권|소프트웨어|개발비|산업재산권', '1038', 'IAS 38 무형자산', 'MEDIUM', '개발비 자산화 요건 엄격, 영업권 비상각→손상검사'),
    (r'보증금|임차보증금', '1116', 'IFRS 16 리스', 'HIGH', '보증금 현재가치 평가, 리스부채와 함께 처리'),
    (r'이연법인세자산', '1012', 'IAS 12 법인세', 'MEDIUM', '항상 비유동, 실현가능성 재평가'),
    (r'투자부동산', '1040', 'IAS 40 투자부동산', 'MEDIUM', '공정가치모형 또는 원가모형 선택'),
    (r'지분법', '1028', 'IAS 28 관계기업투자', 'MEDIUM', '지분법 적용 범위 재검토'),

    # ── 부채 ──
    (r'매입채무', '1109', 'IFRS 9 금융상품', 'LOW', 'AC 측정'),
    (r'단기차입금|장기차입금', '1109', 'IFRS 9 금융상품', 'MEDIUM', '유효이자율법 적용, 공정가치 주석공시'),
    (r'유동성장기', '1001', 'IAS 1 재무제표 표시', 'LOW', '유동/비유동 재분류'),
    (r'사채$', '1109', 'IFRS 9 금융상품', 'MEDIUM', '유효이자율법 적용'),
    (r'전환사채', '1132', 'IAS 32 금융상품 표시', 'HIGH', '부채+자본 분리 (복합금융상품)'),
    (r'미지급금|미지급비용', '1109', 'IFRS 9 금융상품', 'LOW', 'AC 측정'),
    (r'선수금$|선수수익', '1115', 'IFRS 15 수익', 'LOW', '계약부채로 재분류 가능'),
    (r'예수금', '1001', 'IAS 1 재무제표 표시', 'LOW', '표시 변경'),
    (r'퇴직급여.*부채|퇴직급여충당|확정급여', '1019', 'IAS 19 종업원급여', 'HIGH', 'DBO(확정급여채무) 보험수리적 평가 필수'),
    (r'사외적립|퇴직연금운용', '1019', 'IAS 19 종업원급여', 'HIGH', '사외적립자산 공정가치 평가'),
    (r'이연법인세부채', '1012', 'IAS 12 법인세', 'MEDIUM', '항상 비유동 분류'),
    (r'충당부채', '1037', 'IAS 37 충당부채', 'MEDIUM', '인식 요건 재검토'),

    # ── 자본 ──
    (r'자본금$', '1032', 'IAS 32 금융상품 표시', 'LOW', '변경 없음'),
    (r'자본잉여금|주식발행초과금', '1001', 'IAS 1 재무제표 표시', 'LOW', '표시 변경'),
    (r'이익잉여금', '1001', 'IAS 1 재무제표 표시', 'MEDIUM', '전환일 조정사항 반영'),
    (r'기타포괄손익', '1001', 'IAS 1 재무제표 표시', 'MEDIUM', 'FVOCI 평가차이, DBO 재측정 등'),
    (r'자기주식', '1032', 'IAS 32 금융상품 표시', 'LOW', '자본에서 차감 표시'),

    # ── 수익/비용 ──
    (r'매출액|매출$|상품매출|제품매출', '1115', 'IFRS 15 수익', 'HIGH', '5단계 수익인식 모형, 본인/대리인 판단'),
    (r'매출원가', '1002', 'IAS 2 재고자산', 'LOW', 'NRV 평가 영향 가능'),
    (r'급여$|임금$', '1019', 'IAS 19 종업원급여', 'LOW', '변경 없음'),
    (r'임차료|지급임차료', '1116', 'IFRS 16 리스', 'HIGH', '사용권자산 감가상각비 + 리스부채 이자비용으로 대체'),
    (r'감가상각비', '1016', 'IAS 16 유형자산', 'LOW', '내용연수·잔존가치 재검토 반영'),
    (r'대손상각비', '1109', 'IFRS 9 금융상품', 'HIGH', 'ECL 모형으로 대체'),
    (r'이자비용', '1109', 'IFRS 9 금융상품', 'LOW', '유효이자율법 적용'),
    (r'이자수익', '1109', 'IFRS 9 금융상품', 'LOW', '유효이자율법 적용'),
    (r'법인세비용', '1012', 'IAS 12 법인세', 'MEDIUM', '이연법인세 재계산'),
    (r'판매비.*관리비|판관비', '1001', 'IAS 1 재무제표 표시', 'LOW', '기능별/성격별 분류 선택'),
]

# 업종별 영향도 가중치
INDUSTRY_WEIGHT = {
    '도소매업': [(r'리스|임차', 'HIGH'), (r'매출|수익', 'HIGH'), (r'재고', 'HIGH')],
    '제조업': [(r'유형자산|토지|건물|기계', 'HIGH'), (r'차입원가|건설중', 'HIGH')],
    '건설업': [(r'수익|매출', 'HIGH'), (r'공사|진행', 'HIGH')],
    'IT/서비스업': [(r'무형자산|개발비|소프트', 'HIGH'), (r'수익|매출', 'HIGH')],
    '금융업': [(r'금융상품|대출|수취', 'HIGH'), (r'ECL|대손', 'HIGH')],
}

# 필요 내부자료 체크리스트 템플릿
CHECKLIST_TEMPLATE = [
    {'category': '리스 (IFRS 16)', 'items': [
        '전체 임대차 계약서 목록 (건물, 차량, 비품 등)',
        '계약별 월임차료, 계약기간, 보증금 내역',
        '증분차입이자율 산정 근거',
        '변동리스료 해당 여부 (매출연동 등)',
        '원상복구 의무 유무 및 추정 원가',
    ], 'trigger': r'보증금|임차|리스|임대'},
    {'category': '금융상품 (IFRS 9)', 'items': [
        '매출채권 연령분석 (Aging) 자료',
        '과거 3~5년 대손 발생 실적',
        '주요 투자자산의 사업모형 및 현금흐름 특성',
        '금융자산 보유 목적 검토서',
    ], 'trigger': r'매출채권|대손|금융상품|투자증권'},
    {'category': '종업원급여 (IAS 19)', 'items': [
        '임직원 명부 (입사일, 생년월일, 직급, 연봉)',
        '퇴직연금(DB형) 적립 현황',
        '보험수리적 가정 (할인율, 임금상승률, 퇴직률)',
    ], 'trigger': r'퇴직|종업원|확정급여|DBO'},
    {'category': '수익인식 (IFRS 15)', 'items': [
        '주요 수익 계약서 샘플',
        '본인/대리인 판단 근거 (재고위험, 가격결정권)',
        '반품/환불/할인 정책 및 실적',
        '고객충성제도(포인트) 운영 현황',
    ], 'trigger': r'매출|수익|수수료'},
    {'category': '유형자산 (IAS 16)', 'items': [
        '유형자산 대장 (취득원가, 내용연수, 상각방법)',
        '토지/건물 공정가치 평가서 (재평가모형 적용 시)',
        '자산손상 징후 검토',
    ], 'trigger': r'토지|건물|유형자산|기계|차량'},
    {'category': '법인세 (IAS 12)', 'items': [
        '세무조정 내역 (일시적차이 명세)',
        '이월결손금 및 세액공제 현황',
        '적용 법인세율',
    ], 'trigger': r'법인세|이연|세무'},
    {'category': '연결/지분법 (IFRS 10/IAS 28)', 'items': [
        '종속기업·관계기업 목록 및 지분율',
        '종속기업 재무제표',
        '연결범위 판단 근거',
    ], 'trigger': r'지분법|종속|관계기업|연결'},
    {'category': 'IFRS 1 최초적용', 'items': [
        '전환일(개시 재무상태표일) 확정',
        'IFRS 1 면제규정 적용 여부 검토',
        '비교 재무제표 작성 범위',
    ], 'trigger': '.*'},  # 항상 필요
]


# ═══════════════════════════════════════════════════════════
# 컨버전 엔진
# ═══════════════════════════════════════════════════════════

class DartConverter:
    """DART 데이터 기반 K-GAAP → K-IFRS 자동 전환"""

    def __init__(self, api_key: Optional[str] = None):
        self.dart = DartAPI(api_key)
        self.result: Optional[ConversionResult] = None

    def convert(
        self,
        company_name: str,
        fiscal_year: Optional[str] = None,
        fs_div: str = "OFS",
    ) -> ConversionResult:
        """기업명으로 전체 컨버전 파이프라인 실행

        Args:
            company_name: 기업명 (예: "비제바노")
            fiscal_year: 사업연도 (미지정시 직전 연도)
            fs_div: OFS=개별, CFS=연결

        Returns:
            ConversionResult
        """
        if not fiscal_year:
            fiscal_year = str(datetime.now().year - 1)

        print(f"\n{'='*60}")
        print(f"🏢 K-GAAP → K-IFRS 자동 컨버전: {company_name}")
        print(f"{'='*60}")

        # STEP 1: 기업 검색
        print(f"\n[1/7] 기업 검색 중...")
        companies = self.dart.search_company(company_name)
        if not companies:
            raise DartAPIError(f"'{company_name}' 검색 결과가 없습니다.")

        corp = companies[0]
        print(f"  → {corp['corp_name']} (고유번호: {corp['corp_code']})")

        # STEP 2: 기업개황
        print(f"\n[2/7] 기업개황 조회 중...")
        company_info = self.dart.get_company_info(corp['corp_code'])
        if not company_info:
            raise DartAPIError("기업개황 조회 실패")
        print(f"  → 업종: {company_info.industry_category} ({company_info.induty_code})")
        print(f"  → 결산월: {company_info.acc_mt}월, 대표자: {company_info.ceo_nm}")

        # STEP 3: 회계기준 확인
        print(f"\n[3/7] 회계기준 확인 중...")
        acct_std = self.dart.get_accounting_standard(corp['corp_code'], fiscal_year)
        print(f"  → 적용 회계기준: {acct_std}")

        # 결과 객체 초기화
        self.result = ConversionResult(
            company=company_info,
            fiscal_year=fiscal_year,
            accounting_standard=acct_std,
            accuracy_grade="D등급",  # DART 공시 기반
        )

        if acct_std == "K-IFRS":
            self.result.warnings.append(
                f"⚠ {company_name}은 이미 K-IFRS를 적용 중입니다. "
                "전환이 불필요하거나, 개별재무제표(K-GAAP) 기준 전환을 원하시는지 확인하세요."
            )
            print(f"  ⚠ 이미 K-IFRS 적용 중 — 개별재무제표(OFS) 기준으로 진행합니다.")
            fs_div = "OFS"

        # STEP 4: 재무제표 수치 수집
        print(f"\n[4/7] 재무제표 수집 중 (사업연도: {fiscal_year})...")
        all_accounts = self.dart.get_full_financial_statements(
            corp['corp_code'], fiscal_year, fs_div=fs_div
        )

        if not all_accounts:
            # 상세 조회 실패 시 주요계정으로 대체
            print("  → 상세 재무제표 없음, 주요계정으로 대체...")
            all_accounts = self.dart.get_financial_statements(
                corp['corp_code'], fiscal_year, fs_div=fs_div
            )

        if not all_accounts:
            self.result.warnings.append(
                f"⚠ {fiscal_year}년 재무제표 데이터가 없습니다. "
                "사업보고서가 아직 제출되지 않았거나 데이터가 제공되지 않는 기업일 수 있습니다."
            )
            # 전기로 재시도
            prev_year = str(int(fiscal_year) - 1)
            print(f"  → {prev_year}년으로 재시도...")
            all_accounts = self.dart.get_financial_statements(
                corp['corp_code'], prev_year, fs_div=fs_div
            )
            if all_accounts:
                self.result.fiscal_year = prev_year
                fiscal_year = prev_year

        self.result.all_accounts = all_accounts
        self.result.bs_accounts = [a for a in all_accounts if a.sj_div == 'BS']
        self.result.is_accounts = [a for a in all_accounts if a.sj_div in ('IS', 'CIS')]
        self.result.cf_accounts = [a for a in all_accounts if a.sj_div == 'CF']

        print(f"  → BS: {len(self.result.bs_accounts)}개, IS: {len(self.result.is_accounts)}개, CF: {len(self.result.cf_accounts)}개")

        # STEP 5: 감사보고서 검색
        print(f"\n[5/7] 감사보고서 검색 중...")
        audit = self.dart.get_audit_report(corp['corp_code'], fiscal_year)
        if audit:
            self.result.audit_report = audit
            print(f"  → {audit.report_nm} ({audit.rcept_dt})")
        else:
            print("  → 감사보고서를 찾을 수 없습니다.")

        # STEP 6: 전환근거 매핑
        print(f"\n[6/7] K-IFRS 전환근거 매핑 중...")
        self._map_conversion_items(company_info.industry_category)
        print(f"  → {len(self.result.conversion_items)}개 항목 매핑 완료")

        # STEP 7: 체크리스트 생성
        print(f"\n[7/7] 필요 내부자료 체크리스트 생성 중...")
        self._generate_checklist()
        print(f"  → {len(self.result.checklist)}개 카테고리")

        # 전환 변동내역 생성
        self._generate_deltas()

        print(f"\n{'='*60}")
        print(f"✅ 컨버전 완료 (정확도: {self.result.accuracy_grade})")
        print(f"{'='*60}")

        return self.result

    def _map_conversion_items(self, industry: str):
        """재무제표 계정에 K-IFRS 전환근거를 매핑"""
        items = []
        no = 1

        for account in self.result.all_accounts:
            name = account.account_nm
            if not name:
                continue

            # 합계/소계는 스킵
            if re.search(r'총계|합계|소계', name):
                continue

            # K-IFRS 매핑 찾기
            matched = False
            for pattern, kifrs_code, kifrs_name, base_impact, change_desc in ACCOUNT_KIFRS_MAP:
                if re.search(pattern, name):
                    # 업종별 영향도 조정
                    impact = base_impact
                    if industry in INDUSTRY_WEIGHT:
                        for ind_pattern, ind_impact in INDUSTRY_WEIGHT[industry]:
                            if re.search(ind_pattern, name):
                                if self._impact_level(ind_impact) > self._impact_level(impact):
                                    impact = ind_impact

                    # 카테고리 분류
                    cat = '기타'
                    if account.sj_div == 'BS':
                        cat = self._classify_bs_category(name)
                    elif account.sj_div in ('IS', 'CIS'):
                        cat = '수익' if account.ord < 5 else '비용'

                    items.append(ConversionItem(
                        no=no,
                        category=cat,
                        account_name=name,
                        book_value=account.thstrm_amount,
                        kifrs_standard=kifrs_code,
                        kifrs_name=kifrs_name,
                        change_description=change_desc,
                        impact=impact,
                    ))
                    no += 1
                    matched = True
                    break

            if not matched and account.thstrm_amount and abs(account.thstrm_amount) > 0:
                # 매핑 안 된 항목도 기록 (수동 검토 필요)
                items.append(ConversionItem(
                    no=no,
                    category=self._classify_bs_category(name) if account.sj_div == 'BS' else '기타',
                    account_name=name,
                    book_value=account.thstrm_amount,
                    kifrs_standard='-',
                    kifrs_name='수동 검토 필요',
                    change_description='자동 매핑 불가 — 계정과목 확인 후 수동 매핑 필요',
                    impact='REVIEW',
                    accuracy='★☆☆',
                ))
                no += 1

        self.result.conversion_items = items

    def _generate_deltas(self):
        """전환 변동내역 생성"""
        deltas = []
        for item in self.result.conversion_items:
            if item.impact in ('NONE', '-') or item.kifrs_standard == '-':
                changed = False
                adj = 0
            elif item.impact == 'HIGH':
                changed = True
                adj = None  # 실측 필요
            else:
                changed = item.impact == 'MEDIUM'
                adj = None if changed else 0

            kifrs_est = None
            if item.book_value is not None and adj == 0:
                kifrs_est = item.book_value

            deltas.append(ConversionDelta(
                account_name=item.account_name,
                kgaap_amount=item.book_value,
                adjustment=adj,
                kifrs_estimated=kifrs_est,
                changed=changed,
                kifrs_standard=f"K-IFRS {item.kifrs_standard}" if item.kifrs_standard != '-' else '-',
                change_basis=item.change_description,
                assumptions='DART 공시 기반 예비분석. 정밀 산출에는 내부자료 필요.' if changed else '',
            ))

        self.result.conversion_deltas = deltas

    def _generate_checklist(self):
        """필요 내부자료 체크리스트 생성"""
        all_account_names = ' '.join(a.account_nm for a in self.result.all_accounts if a.account_nm)

        checklist = []
        for tmpl in CHECKLIST_TEMPLATE:
            # 해당 카테고리가 이 기업에 관련되는지 확인
            if tmpl['trigger'] == '.*' or re.search(tmpl['trigger'], all_account_names):
                checklist.append({
                    'category': tmpl['category'],
                    'items': tmpl['items'],
                    'priority': '필수' if tmpl['trigger'] == '.*' else '해당시필수',
                })

        self.result.checklist = checklist

    # ── 유틸리티 ──

    @staticmethod
    def _impact_level(impact: str) -> int:
        return {'NONE': 0, 'LOW': 1, 'MEDIUM': 2, 'HIGH': 3, 'REVIEW': 2}.get(impact, 0)

    @staticmethod
    def _classify_bs_category(name: str) -> str:
        """재무상태표 계정의 자산/부채/자본 분류"""
        if re.search(r'자본|잉여금|자기주식|기타포괄', name):
            return '자본'
        if re.search(r'차입금|사채|매입채무|미지급|선수|예수|충당|퇴직|이연법인세부채', name):
            return '부채'
        return '자산'

    # ═══════════════════════════════════════════════════════
    # 엑셀 출력
    # ═══════════════════════════════════════════════════════

    def export_excel(self, output_path: str) -> str:
        """컨버전 결과를 엑셀 파일로 출력"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not self.result:
            raise ValueError("convert()를 먼저 실행하세요.")

        wb = openpyxl.Workbook()
        r = self.result

        # 스타일 정의
        header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1F4E79')
        normal_font = Font(name='맑은 고딕', size=10)
        num_fmt = '#,##0;(#,##0);"-"'
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        # 영향도 색상
        impact_fills = {
            'HIGH': PatternFill('solid', fgColor='FFE0E0'),
            'MEDIUM': PatternFill('solid', fgColor='FFFDE7'),
            'LOW': PatternFill('solid', fgColor='E8F5E9'),
            'NONE': PatternFill('solid', fgColor='F5F5F5'),
            'REVIEW': PatternFill('solid', fgColor='FFE0B2'),
        }

        def write_header(ws, row, headers):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

        def style_data(ws, start_row, end_row, num_cols, num_col_indices=None):
            for row in range(start_row, end_row + 1):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border
                    if num_col_indices and col in num_col_indices:
                        cell.number_format = num_fmt

        # ── Sheet 1: 재무상태표 ──
        ws_bs = wb.active
        ws_bs.title = '재무상태표'
        ws_bs.cell(row=1, column=1, value=f"{r.company.corp_name} 재무상태표 (K-GAAP)").font = Font(name='맑은 고딕', bold=True, size=14)
        ws_bs.cell(row=2, column=1, value=f"제{r.fiscal_year}기 | 정확도: {r.accuracy_grade} | 출처: DART 공시").font = Font(name='맑은 고딕', size=10, color='666666')

        bs_headers = ['계정과목', f'당기금액\n({r.fiscal_year})', '전기금액', '전전기금액']
        write_header(ws_bs, 4, bs_headers)
        row = 5
        for acc in r.bs_accounts:
            ws_bs.cell(row=row, column=1, value=acc.account_nm)
            ws_bs.cell(row=row, column=2, value=acc.thstrm_amount)
            ws_bs.cell(row=row, column=3, value=acc.frmtrm_amount)
            ws_bs.cell(row=row, column=4, value=acc.bfefrmtrm_amount)
            row += 1
        style_data(ws_bs, 5, row - 1, 4, {2, 3, 4})
        ws_bs.column_dimensions['A'].width = 30
        for c in 'BCD':
            ws_bs.column_dimensions[c].width = 18

        # ── Sheet 2: 손익계산서 ──
        ws_is = wb.create_sheet('손익계산서')
        ws_is.cell(row=1, column=1, value=f"{r.company.corp_name} 손익계산서 (K-GAAP)").font = Font(name='맑은 고딕', bold=True, size=14)
        ws_is.cell(row=2, column=1, value=f"제{r.fiscal_year}기 | 출처: DART 공시").font = Font(name='맑은 고딕', size=10, color='666666')

        is_headers = ['계정과목', f'당기금액\n({r.fiscal_year})', '전기금액']
        write_header(ws_is, 4, is_headers)
        row = 5
        for acc in r.is_accounts:
            ws_is.cell(row=row, column=1, value=acc.account_nm)
            ws_is.cell(row=row, column=2, value=acc.thstrm_amount)
            ws_is.cell(row=row, column=3, value=acc.frmtrm_amount)
            row += 1
        style_data(ws_is, 5, row - 1, 3, {2, 3})
        ws_is.column_dimensions['A'].width = 30
        for c in 'BC':
            ws_is.column_dimensions[c].width = 18

        # ── Sheet 3: 전환근거 ──
        ws_conv = wb.create_sheet('전환근거')
        ws_conv.cell(row=1, column=1, value=f"{r.company.corp_name} K-IFRS 전환근거").font = Font(name='맑은 고딕', bold=True, size=14)
        ws_conv.cell(row=2, column=1, value=f"정확도: {r.accuracy_grade} | DART 공시 기반 예비분석 | 정밀 분석에는 내부자료 필요").font = Font(name='맑은 고딕', size=10, color='666666')

        conv_headers = ['No.', '구분', 'K-GAAP 계정과목', '장부금액', '적용 K-IFRS', '전환 시 변동사항', '영향도', '정확도']
        write_header(ws_conv, 4, conv_headers)
        row = 5
        for item in r.conversion_items:
            ws_conv.cell(row=row, column=1, value=item.no)
            ws_conv.cell(row=row, column=2, value=item.category)
            ws_conv.cell(row=row, column=3, value=item.account_name)
            ws_conv.cell(row=row, column=4, value=item.book_value)
            ws_conv.cell(row=row, column=5, value=f"{item.kifrs_standard} {item.kifrs_name}")
            ws_conv.cell(row=row, column=6, value=item.change_description)
            ws_conv.cell(row=row, column=7, value=item.impact)
            ws_conv.cell(row=row, column=8, value=item.accuracy)

            # 영향도 색상
            if item.impact in impact_fills:
                for c in range(1, 9):
                    ws_conv.cell(row=row, column=c).fill = impact_fills[item.impact]
            row += 1

        style_data(ws_conv, 5, row - 1, 8, {4})
        for c, w in zip('ABCDEFGH', [5, 8, 22, 16, 28, 35, 8, 8]):
            ws_conv.column_dimensions[c].width = w

        # ── Sheet 4: 전환 변동내역 ──
        ws_delta = wb.create_sheet('전환변동내역')
        ws_delta.cell(row=1, column=1, value=f"{r.company.corp_name} 전환 변동내역").font = Font(name='맑은 고딕', bold=True, size=14)
        delta_headers = ['계정과목', 'K-GAAP 금액', '전환조정', 'K-IFRS 추정', '변동여부', '적용 K-IFRS', '변동/무변동 근거', '추정 가정']
        write_header(ws_delta, 3, delta_headers)
        row = 4
        for d in r.conversion_deltas:
            ws_delta.cell(row=row, column=1, value=d.account_name)
            ws_delta.cell(row=row, column=2, value=d.kgaap_amount)
            ws_delta.cell(row=row, column=3, value=d.adjustment if d.adjustment is not None else '산출 필요')
            ws_delta.cell(row=row, column=4, value=d.kifrs_estimated if d.kifrs_estimated is not None else '실측 필요')
            ws_delta.cell(row=row, column=5, value='변동' if d.changed else '무변동')
            ws_delta.cell(row=row, column=6, value=d.kifrs_standard)
            ws_delta.cell(row=row, column=7, value=d.change_basis)
            ws_delta.cell(row=row, column=8, value=d.assumptions)
            row += 1
        style_data(ws_delta, 4, row - 1, 8, {2, 3, 4})
        for c, w in zip('ABCDEFGH', [22, 16, 12, 16, 8, 16, 35, 30]):
            ws_delta.column_dimensions[c].width = w

        # ── Sheet 5: 필요 내부자료 ──
        ws_ck = wb.create_sheet('필요내부자료')
        ws_ck.cell(row=1, column=1, value=f"{r.company.corp_name} K-IFRS 전환 필요 내부자료").font = Font(name='맑은 고딕', bold=True, size=14)
        ck_headers = ['카테고리', '필요 자료', '우선순위']
        write_header(ws_ck, 3, ck_headers)
        row = 4
        for cat in r.checklist:
            for i, item in enumerate(cat['items']):
                ws_ck.cell(row=row, column=1, value=cat['category'] if i == 0 else '')
                ws_ck.cell(row=row, column=2, value=item)
                ws_ck.cell(row=row, column=3, value=cat['priority'] if i == 0 else '')
                row += 1
        style_data(ws_ck, 4, row - 1, 3)
        ws_ck.column_dimensions['A'].width = 25
        ws_ck.column_dimensions['B'].width = 55
        ws_ck.column_dimensions['C'].width = 12

        # ── Sheet 6: 기업개요 + 경고 ──
        ws_sum = wb.create_sheet('요약')
        ws_sum.cell(row=1, column=1, value='K-GAAP → K-IFRS 전환 요약').font = Font(name='맑은 고딕', bold=True, size=14)
        info_rows = [
            ('회사명', r.company.corp_name),
            ('사업자등록번호', r.company.bizr_no),
            ('대표이사', r.company.ceo_nm),
            ('업종', f"{r.company.industry_category} ({r.company.induty_code})"),
            ('결산월', f"{r.company.acc_mt}월"),
            ('사업연도', r.fiscal_year),
            ('적용 회계기준', r.accounting_standard),
            ('분석 정확도', r.accuracy_grade),
            ('데이터 출처', 'DART 전자공시시스템'),
            ('분석 일시', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ]
        row = 3
        for label, val in info_rows:
            ws_sum.cell(row=row, column=1, value=label).font = Font(name='맑은 고딕', bold=True, size=10)
            ws_sum.cell(row=row, column=2, value=val).font = normal_font
            row += 1

        if r.warnings:
            row += 1
            ws_sum.cell(row=row, column=1, value='⚠ 주의사항').font = Font(name='맑은 고딕', bold=True, size=11, color='FF0000')
            row += 1
            for w in r.warnings:
                ws_sum.cell(row=row, column=1, value=w).font = normal_font
                row += 1

        row += 1
        ws_sum.cell(row=row, column=1, value='ℹ 정확도 등급 안내').font = Font(name='맑은 고딕', bold=True, size=11)
        row += 1
        grades = [
            'D등급: DART 공시 기반 예비분석. 주요 계정의 전환 방향성 파악 수준.',
            'C등급: 시산표(TB) + 계정과목체계(COA) 입력 시. 계정별 매핑 가능.',
            'B등급: 필수 입력 + 리스/금융상품/임직원 등 추가 입력 시. 주요 항목 실측.',
            'A등급: 전체 입력양식 작성 시. 감사 수준 근접.',
        ]
        for g in grades:
            ws_sum.cell(row=row, column=1, value=g).font = normal_font
            row += 1

        ws_sum.column_dimensions['A'].width = 20
        ws_sum.column_dimensions['B'].width = 50

        # 저장
        wb.save(output_path)
        return output_path


# ═══════════════════════════════════════════════════════════
# CLI 인터페이스
# ═══════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("사용법:")
        print("  python dart_converter.py <기업명> [사업연도]")
        print("")
        print("예시:")
        print("  python dart_converter.py 비제바노 2024")
        print("  python dart_converter.py 금화 2024")
        print("")
        print("API 키 설정:")
        print("  python dart_api.py set-key <발급받은키>")
        sys.exit(1)

    name = sys.argv[1]
    year = sys.argv[2] if len(sys.argv) > 2 else None

    converter = DartConverter()
    result = converter.convert(name, year)

    # 엑셀 출력
    output = f"{name}_KIFRS전환_{result.fiscal_year}.xlsx"
    converter.export_excel(output)
    print(f"\n📊 엑셀 저장 완료: {output}")
