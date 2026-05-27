"""
비정형 엑셀 자동 인식 파서 (K-GAAP → K-IFRS 컨버전용)
=======================================================
어떤 양식의 엑셀이 들어와도 계정과목을 인식하여 정규화된 데이터를 추출한다.
양식(열 위치, 시트명, 헤더 행)이 아니라 **내용(계정과목명, 숫자 패턴)**으로 판단한다.

핵심 원칙:
  1. 시트 유형은 내용물(계정과목 출현빈도, 키워드)로 추론한다
  2. 헤더 행은 키워드 매칭 점수가 가장 높은 행이다  
  3. 열 용도는 헤더 텍스트 + 실제 데이터 타입으로 판단한다
  4. 계정과목은 115+ 패턴으로 인식하고, 인식 못한 과목도 구조적 위치로 추론한다
"""

import re
import openpyxl
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple, Any
from enum import Enum


# ═══════════════════════════════════════════════════════════
# 1. 상수 및 패턴 정의
# ═══════════════════════════════════════════════════════════

class SheetType(Enum):
    """시트 유형"""
    TRIAL_BALANCE = "시산표"
    BALANCE_SHEET = "재무상태표"
    INCOME_STATEMENT = "손익계산서"
    EQUITY_CHANGES = "자본변동표"
    CASH_FLOW = "현금흐름표"
    FIXED_ASSETS = "유형무형자산"
    LEASE = "리스계약"
    EMPLOYEE = "임직원정보"
    BORROWING = "차입금_사채"
    FINANCIAL_INSTRUMENT = "금융상품"
    REVENUE_CONTRACT = "수익계약"
    RELATED_PARTY = "특수관계자"
    TAX = "세무정보"
    COMPANY_PROFILE = "회사개요"
    COA = "계정과목체계"
    CONVERSION_BASIS = "전환근거"
    CONVERSION_DELTA = "전환변동내역"
    CHECKLIST = "체크리스트"
    UNKNOWN = "미분류"


class ColumnRole(Enum):
    """열 용도"""
    ACCOUNT_CODE = "계정코드"
    ACCOUNT_NAME = "계정과목명"
    DEBIT = "차변"
    CREDIT = "대변"
    BALANCE = "잔액"
    AMOUNT_CURRENT = "당기금액"
    AMOUNT_PRIOR = "전기금액"
    AMOUNT_DETAIL = "세부금액"
    AMOUNT_TOTAL = "합계금액"
    KIFRS_CODE = "KIFRS코드"
    KIFRS_NAME = "KIFRS과목명"
    DESCRIPTION = "설명"
    NOTE_REF = "주석참조"
    CATEGORY = "구분"
    PRIORITY = "우선순위"
    IMPACT = "영향도"
    DATE = "날짜"
    NUMBER = "숫자기타"
    TEXT = "텍스트기타"
    ROW_NUMBER = "행번호"
    UNKNOWN = "미분류"


# ── 시트 유형 판별 키워드 (내용 기반) ──
SHEET_TYPE_KEYWORDS = {
    SheetType.TRIAL_BALANCE: {
        'name': ['시산표', 'TB', 'trial balance', '잔액시산', '합계잔액'],
        'content': ['차변', '대변', '차변잔액', '대변잔액', '잔액(차-대)'],
        'weight': 10
    },
    SheetType.BALANCE_SHEET: {
        'name': ['재무상태표', 'BS', 'balance sheet', '대차대조표', 'SOFP'],
        'content': ['유동자산', '비유동자산', '자산총계', '부채총계', '자본총계'],
        'weight': 8
    },
    SheetType.INCOME_STATEMENT: {
        'name': ['손익계산서', 'IS', 'PL', 'income statement', '포괄손익'],
        'content': ['매출액', '매출원가', '영업이익', '당기순이익', '법인세비용'],
        'weight': 8
    },
    SheetType.EQUITY_CHANGES: {
        'name': ['자본변동표', '자본변동', 'equity', 'changes in equity'],
        'content': ['자본금', '자본잉여금', '이익잉여금', '기타포괄'],
        'weight': 8
    },
    SheetType.CASH_FLOW: {
        'name': ['현금흐름표', 'CF', 'cash flow'],
        'content': ['영업활동', '투자활동', '재무활동', '현금및현금성자산'],
        'weight': 8
    },
    SheetType.FIXED_ASSETS: {
        'name': ['유형자산', '무형자산', '자산대장', '고정자산', 'fixed asset'],
        'content': ['취득원가', '감가상각누계액', '장부금액', '내용연수', '상각방법'],
        'weight': 7
    },
    SheetType.LEASE: {
        'name': ['리스', '임대차', 'lease', '임차'],
        'content': ['월임차료', '리스기간', '보증금', '리스개시', '리스종료', '임차료'],
        'weight': 7
    },
    SheetType.EMPLOYEE: {
        'name': ['임직원', '직원', '인사', 'employee', 'HR'],
        'content': ['입사일', '생년월일', '연봉', '퇴직연금', '근속', '직급'],
        'weight': 7
    },
    SheetType.BORROWING: {
        'name': ['차입금', '사채', 'borrowing', 'debt', '금융부채'],
        'content': ['차입종류', '만기일', '이자율', '약정금액', '금융기관'],
        'weight': 7
    },
    SheetType.FINANCIAL_INSTRUMENT: {
        'name': ['금융상품', '매출채권', 'financial', 'ECL'],
        'content': ['연령분석', '미연체', '대손율', '대손충당금', 'ECL', '공정가치'],
        'weight': 7
    },
    SheetType.REVENUE_CONTRACT: {
        'name': ['수익', '매출계약', 'revenue', '수익계약'],
        'content': ['수행의무', '본인/대리인', '총액/순액', '변동대가', '거래처'],
        'weight': 7
    },
    SheetType.RELATED_PARTY: {
        'name': ['특수관계', 'related party', '특관자'],
        'content': ['관계유형', '지분율', '특수관계자명'],
        'weight': 7
    },
    SheetType.TAX: {
        'name': ['세무', 'tax', '법인세', '이연법인세'],
        'content': ['일시적차이', '세무상', '이연법인세', '법인세율', '과세표준'],
        'weight': 7
    },
    SheetType.COMPANY_PROFILE: {
        'name': ['회사개요', '회사정보', 'company', 'profile', '기업개요'],
        'content': ['사업자등록번호', '대표이사', '업종', '설립일', '본점소재지'],
        'weight': 6
    },
    SheetType.COA: {
        'name': ['계정과목', 'COA', 'chart of accounts', '과목체계'],
        'content': ['대분류', '중분류', '재무제표 구분'],
        'weight': 6
    },
    SheetType.CONVERSION_BASIS: {
        'name': ['전환근거', '매핑', 'conversion', 'mapping'],
        'content': ['적용 K-IFRS', '영향도', '변동사항', '전환 시'],
        'weight': 5
    },
    SheetType.CONVERSION_DELTA: {
        'name': ['변동내역', '전환조정', 'delta', 'adjustment'],
        'content': ['K-GAAP', 'K-IFRS', '전환 조정', '변동 여부', '추정'],
        'weight': 5
    },
    SheetType.CHECKLIST: {
        'name': ['체크리스트', 'checklist', '필요자료', '내부자료', '요청자료', '수집자료'],
        'content': ['완료 여부', '담당부서', '수집 현황', '제출', '요청', '필수', '선택',
                     '완료', '미완료', '검토 필요', '자료명', '우선순위'],
        'weight': 8
    },
}

# ── 네거티브 키워드: 이 키워드가 있으면 해당 SheetType에서 점수 차감 ──
SHEET_TYPE_NEGATIVE_KW = {
    SheetType.LEASE: ['체크리스트', '필요자료', '내부자료', '전환근거', '변동내역',
                      '완료 여부', '담당부서'],
    SheetType.REVENUE_CONTRACT: ['체크리스트', '필요자료', '내부자료', '전환근거',
                                  '완료 여부', '담당부서', '우선순위'],
    SheetType.FINANCIAL_INSTRUMENT: ['체크리스트', '필요자료', '내부자료'],
    SheetType.EMPLOYEE: ['체크리스트', '필요자료', '내부자료'],
}

# ── 열 용도 판별 키워드 ──
COLUMN_KEYWORDS = {
    ColumnRole.ACCOUNT_CODE: ['계정코드', '코드', 'code', '계정번호', 'acct_cd', '과목코드'],
    ColumnRole.ACCOUNT_NAME: ['계정과목', '과목명', '계정명', 'account', '과 목', '과  목', '계정과목명', '항목'],
    ColumnRole.DEBIT: ['차변', '차변잔액', 'debit', '차변합계'],
    ColumnRole.CREDIT: ['대변', '대변잔액', 'credit', '대변합계'],
    ColumnRole.BALANCE: ['잔액', '기말잔액', 'balance', '잔액(차-대)', '기말'],
    ColumnRole.AMOUNT_CURRENT: ['당기', '당 기', '제.*당.*기', '금액', '금 액'],
    ColumnRole.AMOUNT_PRIOR: ['전기', '전 기', '제.*전.*기', '비교'],
    ColumnRole.AMOUNT_DETAIL: ['세부', '세부금액'],
    ColumnRole.AMOUNT_TOTAL: ['합계', '합 계', '총계', 'total', '소계'],
    ColumnRole.KIFRS_CODE: ['K-IFRS.*코드', 'KIFRS.*코드', 'IFRS.*매핑', '매핑코드'],
    ColumnRole.KIFRS_NAME: ['K-IFRS.*과목', 'KIFRS.*과목', 'IFRS.*계정'],
    ColumnRole.DESCRIPTION: ['설명', '비고', '내용', '근거', 'description', 'note', '유의사항'],
    ColumnRole.NOTE_REF: ['주석', '참조', 'ref'],
    ColumnRole.CATEGORY: ['구분', '분류', 'category', '영역'],
    ColumnRole.PRIORITY: ['우선순위', 'priority', '중요도'],
    ColumnRole.IMPACT: ['영향도', 'impact', '영향'],
    ColumnRole.ROW_NUMBER: ['no', 'no.', '번호', 'seq', '#'],
}

# ── 계정과목 인식 패턴 (auto_mapping_logic.md 기반 + 확장) ──
ACCOUNT_PATTERNS = {
    # ── 자산 ──
    r'현금|현금성': '현금및현금성자산',
    r'단기금융상품|정기예금|적금': '단기금융상품',
    r'매출채권(?!.*연령)': '매출채권',
    r'대손충당금': '대손충당금',
    r'미수금': '미수금',
    r'미수수익': '미수수익',
    r'선급금': '선급금',
    r'선급비용': '선급비용',
    r'^재고자산$|^상품$|^제품$|^원재료$|^재공품$|^저장품$': '재고자산',
    r'장기금융상품': '장기금융상품',
    r'매도가능증권|투자증권|지분증권': '투자증권',
    r'장기대여금|단기대여금|^대여금$': '대여금',
    r'^토지$': '토지',
    r'^건물$': '건물',
    r'차량': '차량운반구',
    r'기계장치|^설비$': '기계장치',
    r'비품|집기|공구': '비품',
    r'건설중인자산': '건설중인자산',
    r'무형자산|^영업권$|소프트웨어|^개발비$|산업재산권': '무형자산',
    r'보증금|임차보증금': '보증금',
    r'이연법인세자산': '이연법인세자산',
    r'투자부동산': '투자부동산',
    r'장기매출채권': '장기매출채권',
    r'지점장치|매장시설': '매장시설',
    r'당기손익.*공정가치.*금융자산|FVTPL.*금융자산': 'FVTPL금융자산',
    r'기타포괄.*공정가치.*금융자산|FVOCI.*금융자산': 'FVOCI금융자산',
    r'상각후원가.*금융자산|AC금융자산': 'AC금융자산',
    r'사용권자산': '사용권자산',
    r'리스부채': '리스부채',
    r'순확정급여.*부채|순확정급여': '순확정급여부채',
    r'기타수취채권|기타채권': '기타수취채권',
    r'기타지급채무|기타채무': '기타지급채무',
    r'계약자산': '계약자산',
    r'계약부채': '계약부채',
    r'기타유동자산$': '기타유동자산',
    r'기타비유동자산$': '기타비유동자산',
    r'기타유동부채$': '기타유동부채',
    r'기타비유동부채$': '기타비유동부채',
    r'지분법.*투자주식|지분법적용투자': '지분법적용투자주식',
    r'^장기금융자산$': '장기금융자산',
    r'파생금융.*부채|파생상품부채': '파생금융부채',
    r'파생금융.*자산|파생상품자산': '파생금융자산',
    r'환불충당부채': '환불충당부채',
    r'미지급법인세|법인세부채': '미지급법인세',
    r'선급법인세|법인세자산': '선급법인세',
    r'기타.*유형자산|기타유형자산': '기타유형자산',
    r'^단기투자증권$': '단기투자증권',
    r'상품평가충당금': '상품평가충당금',
    r'원재료평가충당금': '원재료평가충당금',
    r'^원재료$': '원재료',
    r'^저장품$': '저장품',
    r'^재공품$': '재공품',
    r'감가상각누계액': '감가상각누계액',
    r'유형자산.*순액|유형자산합계': '유형자산(순액)',
    r'당좌자산': '당좌자산',
    r'투자자산$|투자자산소계': '투자자산(소계)',
    r'유동성사채': '유동성사채',
    r'재고자산평가손실': '재고자산평가손실',

    # ── 부채 ──
    r'매입채무': '매입채무',
    r'^단기차입금$': '단기차입금',
    r'^장기차입금$': '장기차입금',
    r'유동성장기부채|유동성장기': '유동성장기부채',
    r'^사채$': '사채',
    r'전환사채': '전환사채',
    r'^미지급금$': '미지급금',
    r'^선수금$': '선수금',
    r'선수수익': '선수수익',
    r'^예수금$': '예수금',
    r'미지급비용': '미지급비용',
    r'수입보증금|임대보증금': '수입보증금',
    r'퇴직급여충당|퇴직급여부채|확정급여': '퇴직급여부채',
    r'퇴직연금운용자산|사외적립자산': '사외적립자산',
    r'이연법인세부채': '이연법인세부채',
    r'^충당부채$': '충당부채',

    # ── 자본 ──
    r'^자본금$': '자본금',
    r'자본잉여금|주식발행초과금': '자본잉여금',
    r'이익잉여금': '이익잉여금',
    r'기타포괄손익누계': '기타포괄손익누계액',
    r'자기주식': '자기주식',
    r'^자본조정$': '자본조정',

    # ── 손익: 매출/수익 ──
    r'^매출액$|^매출$|^상품매출$|^제품매출$': '매출액',
    r'수수료매출|용역매출|서비스매출': '수수료매출',
    r'임대수익|임대료수입': '임대수익',
    r'관리비수익': '관리비수익',
    r'^매출원가$': '매출원가',
    r'^상품매출원가$': '상품매출원가',

    # ── 손익: 판관비 세부 ──
    r'^급여$|^임금$|^급료$': '급여',
    r'퇴직급여(?!.*충당|.*부채)': '퇴직급여(비용)',
    r'복리후생비': '복리후생비',
    r'여비교통비|출장비': '여비교통비',
    r'^통신비$': '통신비',
    r'^수도광열비$|^전기료$|전력비': '수도광열비',
    r'세금과공과|공과금': '세금과공과',
    r'임차료|지급임차료': '임차료',
    r'^감가상각비$': '감가상각비',
    r'무형자산상각': '무형자산상각비',
    r'^수선비$|수선유지비': '수선비',
    r'^보험료$': '보험료',
    r'^접대비$': '접대비',
    r'광고선전비|광고비|선전비': '광고선전비',
    r'운반비|배달비|배송비': '운반비',
    r'지급수수료|수수료비용': '지급수수료',
    r'대손상각비': '대손상각비',
    r'^소모품비$': '소모품비',
    r'차량유지비|차량비': '차량유지비',
    r'^교육훈련비$|교육비': '교육훈련비',
    r'^도서인쇄비$|인쇄비|도서비': '도서인쇄비',
    r'잡비|잡손실': '잡비',
    r'경상연구개발비|연구비': '경상연구개발비',
    r'포장비': '포장비',
    r'^판매촉진비$': '판매촉진비',
    r'^견본비$': '견본비',
    r'^위탁판매수수료$': '위탁판매수수료',
    r'하역비|상하차비': '하역비',
    r'^잡급$': '잡급',
    r'상여금|상여': '상여금',
    r'^외주가공비$|외주비': '외주가공비',
    r'판매비.*관리비|판관비': '판매비와관리비',

    # ── 손익: 영업외 ──
    r'^이자비용$': '이자비용',
    r'^이자수익$': '이자수익',
    r'배당금수익|수입배당금': '배당금수익',
    r'외환차익': '외환차익',
    r'외환차손': '외환차손',
    r'외화환산이익': '외화환산이익',
    r'외화환산손실': '외화환산손실',
    r'유형자산처분이익|유형자산처분': '유형자산처분이익',
    r'유형자산처분손실': '유형자산처분손실',
    r'투자자산처분이익|투자자산처분': '투자자산처분이익',
    r'투자자산처분손실': '투자자산처분손실',
    r'지분법.*이익|지분법이익': '지분법이익',
    r'지분법.*손실|지분법손실': '지분법손실',
    r'잡이익|잡수입': '잡이익',
    r'기부금': '기부금',
    r'법인세비용': '법인세비용',
    r'^영업이익$|^영업손실$': '영업이익',
    r'^영업외수익$': '영업외수익',
    r'^영업외비용$': '영업외비용',
    r'당기순이익|당기순손실': '당기순이익',
    r'총포괄손익|총포괄이익': '총포괄손익',
    r'^매출총이익$|매출총손실': '매출총이익',
    r'^영업손익$': '영업손익',
    r'^기타수익$': '기타수익',
    r'^기타비용$': '기타비용',
    r'^금융수익$': '금융수익',
    r'^금융비용$': '금융비용',
    r'법인세.*차감.*이익|법인세차감전': '법인세차감전순이익',
    r'중단사업.*손익': '중단사업손익',
    r'계속사업.*손익|계속사업이익': '계속사업손익',
    r'상품매출액': '상품매출액',
    r'기타매출|기타매출액': '기타매출액',
    r'임대매출|임대매출액': '임대매출액',
    r'^상품매출원가$': '상품매출원가',
    r'임대원가': '임대원가',
    r'기초상품재고|기초재고': '기초상품재고액',
    r'당기상품매입|당기매입': '당기상품매입액',
    r'기말상품재고|기말재고': '기말상품재고액',
    r'타계정.*대체': '타계정대체',
    r'수선수입': '수선수입',
    r'수수료수익|수입수수료': '수수료수익',
    r'경상개발비': '경상개발비',
    r'주당.*손익|주당이익|주당순이익|EPS': '주당손익',
    r'재고자산평가.*환입': '재고자산평가충당금환입',

    # ── 재무상태표 합계 ──
    r'유동자산$|Ⅰ.*유동자산': '유동자산(소계)',
    r'비유동자산$|Ⅱ.*비유동자산': '비유동자산(소계)',
    r'자산총계|자산 총계|자산합계': '자산총계',
    r'유동부채$|Ⅰ.*유동부채': '유동부채(소계)',
    r'비유동부채$|Ⅱ.*비유동부채': '비유동부채(소계)',
    r'부채총계|부채 총계|부채합계': '부채총계',
    r'자본총계|자본 총계|자본합계': '자본총계',
    r'부채.*자본.*총계': '부채와자본총계',

    # ── 현금흐름 ──
    r'영업활동.*현금|영업활동으로': '영업활동현금흐름',
    r'투자활동.*현금|투자활동으로': '투자활동현금흐름',
    r'재무활동.*현금|재무활동으로': '재무활동현금흐름',
    r'현금.*증감|현금의증가': '현금의증감',
    r'기초.*현금|기초현금': '기초현금',
    r'기말.*현금|기말현금': '기말현금',
    r'재고자산.*감소|재고자산.*증가': '재고자산증감',
    r'매출채권.*감소|매출채권.*증가': '매출채권증감',
    r'매입채무.*증가|매입채무.*감소': '매입채무증감',
    r'미지급금.*증가|미지급금.*감소': '미지급금증감',
    r'선수금.*증가|선수금.*감소': '선수금증감',
    r'예수금.*증가|예수금.*감소': '예수금증감',
    r'퇴직금.*지급': '퇴직금지급',
    r'퇴직연금.*감소|퇴직연금.*증가|사외적립자산.*감소': '퇴직연금운용자산증감',
    r'사채할인발행차금': '사채할인발행차금',
}

# 합계/소계 패턴 (들여쓰기 구조 파악용)
SUBTOTAL_PATTERNS = re.compile(
    r'합계|총계|소계|계$|Ⅰ\.|Ⅱ\.|Ⅲ\.|Ⅳ\.|Ⅴ\.'
)


# ═══════════════════════════════════════════════════════════
# 2. 데이터 구조
# ═══════════════════════════════════════════════════════════

@dataclass
class ColumnInfo:
    """열 정보"""
    index: int                     # 1-based column index
    letter: str                    # 'A', 'B', ...
    role: ColumnRole = ColumnRole.UNKNOWN
    header_text: str = ""          # 원본 헤더 텍스트
    confidence: float = 0.0        # 판별 신뢰도 (0~1)
    sample_values: list = field(default_factory=list)

@dataclass
class ParsedRow:
    """파싱된 행"""
    row_number: int
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    recognized_as: Optional[str] = None      # 표준 계정과목명
    amounts: Dict[str, Any] = field(default_factory=dict)  # role → value
    is_subtotal: bool = False
    is_header: bool = False
    is_category: bool = False                # "자 산", "부 채" 등 분류행
    indent_level: int = 0                    # 들여쓰기 수준
    raw_values: Dict[int, Any] = field(default_factory=dict)

@dataclass
class ParsedSheet:
    """파싱된 시트"""
    name: str
    sheet_type: SheetType
    type_confidence: float
    header_row: int
    columns: List[ColumnInfo] = field(default_factory=list)
    rows: List[ParsedRow] = field(default_factory=list)
    metadata: Dict[str, str] = field(default_factory=dict)  # 회사명, 기간 등

@dataclass
class ParsedWorkbook:
    """파싱된 워크북"""
    filename: str
    sheets: List[ParsedSheet] = field(default_factory=list)
    company_name: Optional[str] = None
    fiscal_period: Optional[str] = None


# ═══════════════════════════════════════════════════════════
# 3. 핵심 엔진
# ═══════════════════════════════════════════════════════════

class ExcelParser:
    """비정형 엑셀 자동 인식 파서"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.result = ParsedWorkbook(filename=filepath)

    def parse(self) -> ParsedWorkbook:
        """전체 워크북 파싱"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            parsed = self._parse_sheet(ws, sheet_name)
            self.result.sheets.append(parsed)

        # 워크북 수준 메타데이터 추출
        self._extract_workbook_metadata()
        return self.result

    def _parse_sheet(self, ws, sheet_name: str) -> ParsedSheet:
        """단일 시트 파싱"""
        # Step 1: 시트 유형 추론
        sheet_type, type_conf = self._classify_sheet(ws, sheet_name)

        # Step 2: 헤더 행 탐지
        header_row = self._find_header_row(ws)

        # Step 3: 열 분류
        columns = self._classify_columns(ws, header_row, sheet_type)

        # Step 4: 데이터 행 파싱
        rows = self._parse_data_rows(ws, header_row, columns, sheet_type)

        # Step 5: 메타데이터 추출 (헤더 위 행들에서)
        metadata = self._extract_sheet_metadata(ws, header_row)

        return ParsedSheet(
            name=sheet_name,
            sheet_type=sheet_type,
            type_confidence=type_conf,
            header_row=header_row,
            columns=columns,
            rows=rows,
            metadata=metadata,
        )

    # ─────────────────────────────────────────────
    # 3-1. 시트 유형 분류 (이름 + 내용 기반)
    # ─────────────────────────────────────────────
    def _classify_sheet(self, ws, sheet_name: str) -> Tuple[SheetType, float]:
        """시트 유형을 이름과 내용으로 추론"""
        scores = {}
        name_lower = sheet_name.lower().replace(' ', '')

        # 시트 내용 샘플링 (처음 20행)
        content_texts = []
        for r in range(1, min(21, (ws.max_row or 1) + 1)):
            for c in range(1, min(15, (ws.max_column or 1) + 1)):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and len(v.strip()) > 0:
                    content_texts.append(v.strip())
        all_content = ' '.join(content_texts).lower()

        for stype, config in SHEET_TYPE_KEYWORDS.items():
            score = 0.0
            # 시트명 매칭 (가중치 높음)
            for kw in config['name']:
                if kw.lower() in name_lower:
                    score += config['weight'] * 2
                    break
            # 내용 매칭
            content_hits = 0
            for kw in config['content']:
                if kw.lower() in all_content:
                    content_hits += 1
            if config['content']:
                score += (content_hits / len(config['content'])) * config['weight']
            # 네거티브 키워드 차감
            if stype in SHEET_TYPE_NEGATIVE_KW:
                neg_hits = 0
                for nkw in SHEET_TYPE_NEGATIVE_KW[stype]:
                    if nkw.lower() in all_content or nkw.lower() in name_lower:
                        neg_hits += 1
                score -= neg_hits * 3  # 네거티브 키워드 1개당 -3점
            scores[stype] = max(score, 0.0)

        if not scores or max(scores.values()) == 0:
            return SheetType.UNKNOWN, 0.0

        best = max(scores, key=scores.get)
        max_score = scores[best]
        # 정규화된 신뢰도
        confidence = min(max_score / 20.0, 1.0)
        return best, confidence

    # ─────────────────────────────────────────────
    # 3-2. 헤더 행 탐지 (키워드 점수 기반, 병합셀 대응)
    # ─────────────────────────────────────────────
    def _find_header_row(self, ws) -> int:
        """키워드 매칭 점수가 가장 높은 행을 헤더로 판정

        개선사항:
        - 병합셀 영역에 속한 행도 병합 기준셀 값을 읽음
        - 헤더 전용 고가중치 키워드 분리 (과목, 금액, 당기, 전기 등)
        - 긴 텍스트(설명/데이터행)에 패널티 부여
        - 전체 행 병합(타이틀 행) 자동 제외
        """
        best_row = 1
        best_score = -1
        max_check = min(20, (ws.max_row or 1))

        # 병합셀 맵 구축: (row, col) → 실제 값이 있는 셀
        merged_map = {}
        for mr in ws.merged_cells.ranges:
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    merged_map[(row, col)] = (mr.min_row, mr.min_col)

        def get_cell_value(row, col):
            """병합셀이면 기준셀 값을 반환"""
            if (row, col) in merged_map:
                src_row, src_col = merged_map[(row, col)]
                return ws.cell(row=src_row, column=src_col).value
            return ws.cell(row=row, column=col).value

        # 전체 폭 병합 행(타이틀/구분행) 식별 → 제외 대상
        title_rows = set()
        max_col = min(20, (ws.max_column or 1))
        for mr in ws.merged_cells.ranges:
            span = mr.max_col - mr.min_col + 1
            if span >= max_col * 0.7:  # 전체 열의 70% 이상 병합 → 타이틀
                for row in range(mr.min_row, mr.max_row + 1):
                    title_rows.add(row)

        # 헤더 전용 고가중치 키워드 (열 이름에만 나타나는 용어)
        HEADER_STRONG_KW = {
            '과목', '항목', 'account', '당기', '전기', '차변', '대변',
            '잔액', '금액', '코드', '구분', '합계', '비고', '설명',
            '영향도', '우선순위', 'no.', '번호',
        }
        # 일반 키워드 (본문에도 나올 수 있음)
        all_keywords = set()
        for keywords in COLUMN_KEYWORDS.values():
            all_keywords.update(kw.lower() for kw in keywords)

        for r in range(1, max_check + 1):
            if r in title_rows:
                continue

            score = 0
            text_cells = 0
            num_cells = 0
            total_text_len = 0

            for c in range(1, max_col + 1):
                v = get_cell_value(r, c)
                if v is None:
                    continue
                if isinstance(v, str):
                    # 멀티라인 텍스트에서 첫 줄만 헤더 판별에 사용
                    first_line = v.split('\n')[0].strip()
                    text_cells += 1
                    total_text_len += len(first_line)
                    v_clean = first_line.lower().replace(' ', '')

                    # 고가중치 헤더 키워드 매칭
                    for kw in HEADER_STRONG_KW:
                        if kw in v_clean:
                            score += 8
                    # 일반 키워드 매칭
                    for kw in all_keywords:
                        if kw in v_clean:
                            score += 2
                elif isinstance(v, (int, float)):
                    num_cells += 1

            # 헤더 행은 텍스트 비율이 높고 숫자가 적어야 함
            if text_cells > 0 and text_cells >= num_cells:
                score += text_cells * 2
            else:
                score *= 0.2  # 숫자가 많으면 데이터 행일 가능성 높음

            # 긴 텍스트 패널티: 헤더 셀은 보통 짧음 (평균 20자 이내)
            if text_cells > 0:
                avg_len = total_text_len / text_cells
                if avg_len > 30:
                    score *= 0.4  # 설명/데이터 행
                elif avg_len > 15:
                    score *= 0.8

            if score > best_score:
                best_score = score
                best_row = r

        return best_row

    # ─────────────────────────────────────────────
    # 3-3. 열 분류 (헤더 텍스트 + 데이터 타입)
    # ─────────────────────────────────────────────
    def _classify_columns(self, ws, header_row: int, sheet_type: SheetType) -> List[ColumnInfo]:
        """각 열의 용도를 판별"""
        columns = []
        max_col = ws.max_column or 1

        for c in range(1, max_col + 1):
            header_val = ws.cell(row=header_row, column=c).value
            header_text = str(header_val).strip() if header_val else ""
            header_clean = header_text.lower().replace(' ', '').replace('\n', '')

            # 샘플 데이터 수집 (헤더 아래 10행)
            samples = []
            for r in range(header_row + 1, min(header_row + 11, (ws.max_row or header_row) + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    samples.append(v)

            # 키워드 매칭으로 열 용도 판별
            best_role = ColumnRole.UNKNOWN
            best_conf = 0.0

            for role, keywords in COLUMN_KEYWORDS.items():
                for kw in keywords:
                    kw_clean = kw.lower().replace(' ', '')
                    if re.search(kw_clean, header_clean):
                        match_conf = 0.9
                        if best_conf < match_conf:
                            best_role = role
                            best_conf = match_conf

            # 키워드로 판별 안 되면 데이터 타입으로 추론
            if best_role == ColumnRole.UNKNOWN and samples:
                num_count = sum(1 for s in samples if isinstance(s, (int, float)))
                str_count = sum(1 for s in samples if isinstance(s, str))

                if num_count > len(samples) * 0.7:
                    # 숫자 열 → 금액일 가능성
                    if any(abs(s) > 1000 for s in samples if isinstance(s, (int, float))):
                        best_role = ColumnRole.NUMBER
                        best_conf = 0.3
                elif str_count > len(samples) * 0.7:
                    # 문자 열 → 계정과목 탐지
                    account_hits = sum(1 for s in samples if isinstance(s, str) and self._is_account_name(s))
                    if account_hits > len(samples) * 0.3:
                        best_role = ColumnRole.ACCOUNT_NAME
                        best_conf = 0.7
                    else:
                        best_role = ColumnRole.TEXT
                        best_conf = 0.3

            # 행번호 열 판별 (1, 2, 3... 순서인 경우)
            if best_role in (ColumnRole.NUMBER, ColumnRole.UNKNOWN) and samples:
                if all(isinstance(s, (int, float)) for s in samples[:5]):
                    vals = [int(s) for s in samples[:5] if isinstance(s, (int, float))]
                    if vals == list(range(vals[0], vals[0] + len(vals))):
                        best_role = ColumnRole.ROW_NUMBER
                        best_conf = 0.8

            columns.append(ColumnInfo(
                index=c,
                letter=get_column_letter(c),
                role=best_role,
                header_text=header_text,
                confidence=best_conf,
                sample_values=samples[:5],
            ))

        # 후처리: 금액 열이 여러 개일 때 당기/전기 구분
        self._disambiguate_amount_columns(columns, ws, header_row, sheet_type)
        return columns

    def _disambiguate_amount_columns(self, columns: List[ColumnInfo], ws, header_row: int, sheet_type: SheetType):
        """숫자 열이 여러 개일 때 당기/전기/세부/합계를 구분"""
        num_cols = [c for c in columns if c.role in (ColumnRole.NUMBER, ColumnRole.UNKNOWN)
                    and any(isinstance(v, (int, float)) for v in c.sample_values)]

        if len(num_cols) < 2:
            return

        # 재무제표 유형이면: 보통 세부|합계|세부|합계 또는 당기|전기 패턴
        if sheet_type in (SheetType.BALANCE_SHEET, SheetType.INCOME_STATEMENT):
            if len(num_cols) == 2:
                num_cols[0].role = ColumnRole.AMOUNT_CURRENT
                num_cols[0].confidence = 0.6
                num_cols[1].role = ColumnRole.AMOUNT_PRIOR
                num_cols[1].confidence = 0.6
            elif len(num_cols) == 4:
                num_cols[0].role = ColumnRole.AMOUNT_DETAIL
                num_cols[1].role = ColumnRole.AMOUNT_CURRENT
                num_cols[2].role = ColumnRole.AMOUNT_DETAIL
                num_cols[3].role = ColumnRole.AMOUNT_PRIOR
                for nc in num_cols:
                    nc.confidence = 0.5

        # 시산표면: 차변|대변|잔액 패턴
        elif sheet_type == SheetType.TRIAL_BALANCE:
            assigned = {c.role for c in columns if c.role in (ColumnRole.DEBIT, ColumnRole.CREDIT, ColumnRole.BALANCE)}
            unassigned = [c for c in num_cols if c.role not in assigned]
            roles_needed = [r for r in (ColumnRole.DEBIT, ColumnRole.CREDIT, ColumnRole.BALANCE) if r not in assigned]
            for col, role in zip(unassigned, roles_needed):
                col.role = role
                col.confidence = 0.5

    # ─────────────────────────────────────────────
    # 3-4. 데이터 행 파싱
    # ─────────────────────────────────────────────
    def _parse_data_rows(self, ws, header_row: int, columns: List[ColumnInfo], sheet_type: SheetType) -> List[ParsedRow]:
        """데이터 행을 파싱하여 정규화"""
        rows = []
        name_col = next((c for c in columns if c.role == ColumnRole.ACCOUNT_NAME), None)
        code_col = next((c for c in columns if c.role == ColumnRole.ACCOUNT_CODE), None)

        # 계정과목 열이 없으면 첫 번째 텍스트 열을 사용
        if not name_col:
            for c in columns:
                if c.role in (ColumnRole.TEXT, ColumnRole.UNKNOWN):
                    if any(isinstance(v, str) for v in c.sample_values):
                        name_col = c
                        name_col.role = ColumnRole.ACCOUNT_NAME
                        name_col.confidence = 0.4
                        break
            # 그래도 없으면 첫 번째 열
            if not name_col and columns:
                name_col = columns[0]

        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            # 빈 행 스킵
            all_none = all(ws.cell(row=r, column=c.index).value is None for c in columns)
            if all_none:
                continue

            parsed = ParsedRow(row_number=r, raw_values={})

            # 각 열의 값 추출
            for col in columns:
                val = ws.cell(row=r, column=col.index).value
                parsed.raw_values[col.index] = val

                if col.role == ColumnRole.ACCOUNT_CODE and val is not None:
                    parsed.account_code = str(val).strip()
                elif col == name_col and val is not None:
                    raw_name = str(val).strip()
                    # 공백, 특수문자 정리
                    clean_name = re.sub(r'\s{2,}', ' ', raw_name).strip()
                    parsed.account_name = clean_name

                    # 들여쓰기 수준 판단
                    leading_spaces = len(raw_name) - len(raw_name.lstrip())
                    parsed.indent_level = leading_spaces // 2

                    # 계정과목 인식
                    parsed.recognized_as = self._recognize_account(clean_name)

                    # 분류행 판단 ("자 산", "부 채", "자 본" 등)
                    if re.match(r'^[자부자손수비][  ]*[산채본익출용]', raw_name):
                        parsed.is_category = True

                    # 소계/합계 판단
                    if SUBTOTAL_PATTERNS.search(clean_name):
                        parsed.is_subtotal = True

                elif col.role in (ColumnRole.DEBIT, ColumnRole.CREDIT, ColumnRole.BALANCE,
                                  ColumnRole.AMOUNT_CURRENT, ColumnRole.AMOUNT_PRIOR,
                                  ColumnRole.AMOUNT_DETAIL, ColumnRole.AMOUNT_TOTAL,
                                  ColumnRole.NUMBER):
                    if isinstance(val, (int, float)):
                        parsed.amounts[col.role.value] = val

            # 계정과목명이 있는 행만 추가 (완전히 빈 행 제외)
            if parsed.account_name or parsed.account_code or parsed.amounts:
                rows.append(parsed)

        return rows

    # ─────────────────────────────────────────────
    # 3-5. 계정과목 인식
    # ─────────────────────────────────────────────
    def _recognize_account(self, name: str) -> Optional[str]:
        """계정과목명을 표준 계정과목으로 인식

        개선사항:
        - "미수금 / 미수수익 / 단기대여금" → 첫 번째 인식 가능한 과목 반환
        - "(구: 단기투자증권)" 같은 괄호 설명 제거
        - 주석참조, K-IFRS 변환 설명 등 노이즈 제거
        """
        if not name:
            return None

        # 멀티라인 셀 처리: 첫 줄만 계정과목 인식에 사용
        if '\n' in name:
            first_line = name.split('\n')[0].strip()
            result = self._recognize_account(first_line)
            if result:
                return result
            # 첫 줄 실패시 줄바꿈 제거 후 전체로 시도

        # 전처리: 노이즈 제거
        clean = re.sub(r'\(주석\s*[\d,\s]+\)', '', name)   # "(주석 2,10)" 제거
        clean = re.sub(r'\(주석\d+\)', '', clean)           # "(주석3)" 제거
        clean = re.sub(r'\(주\s*\d+\)', '', clean)          # "(주 3)" 제거
        clean = re.sub(r'\(구:.*?\)', '', clean)            # "(구: 단기투자증권)" 제거
        clean = re.sub(r'\(순액\)', '', clean)              # "(순액)" 제거
        clean = re.sub(r'\(.*?조정.*?\)', '', clean)        # "(반품조정 포함)" 등 제거
        # 선행 번호 제거: "1.", "(1)", "I.", "II.", "가.", "나." 등
        clean = re.sub(r'^[\s]*(?:\(?\d+\)?\.?\s*|[IVⅠⅡⅢⅣⅤivx]+\.\s*|[가-힣]\.?\s+)', '', clean)
        # 주석 참조 패턴: "~(주석 3,8,17)" 또는 "~(주석2,4)"
        clean = re.sub(r'\(주석[\s\d,]+\)', '', clean)
        clean = clean.strip()

        # 슬래시/쉼표로 구분된 다중 계정 처리
        # "미수금 / 미수수익 / 단기대여금" → ['미수금', '미수수익', '단기대여금']
        if '/' in clean or ',' in clean:
            parts = re.split(r'\s*/\s*|\s*,\s*', clean)
            for part in parts:
                result = self._match_single_account(part.strip())
                if result:
                    return result
            # 전체 문자열로도 재시도
            return self._match_single_account(re.sub(r'\s+', '', clean))

        return self._match_single_account(clean)

    def _match_single_account(self, name: str) -> Optional[str]:
        """단일 계정과목명을 패턴과 매칭"""
        if not name:
            return None

        clean = re.sub(r'\s+', '', name)  # 공백 제거
        clean = clean.strip()

        if len(clean) < 2:
            return None

        # 1순위: 정확 매칭
        for pattern, standard_name in ACCOUNT_PATTERNS.items():
            if re.fullmatch(pattern, clean):
                return standard_name

        # 2순위: 포함 매칭 (긴 패턴 우선)
        matches = []
        for pattern, standard_name in ACCOUNT_PATTERNS.items():
            if re.search(pattern, clean):
                matches.append((len(pattern), standard_name))
        if matches:
            matches.sort(key=lambda x: -x[0])  # 긴 패턴 우선
            return matches[0][1]

        return None

    def _is_account_name(self, text: str) -> bool:
        """텍스트가 계정과목명인지 판단"""
        return self._recognize_account(text) is not None

    # ─────────────────────────────────────────────
    # 3-6. 메타데이터 추출
    # ─────────────────────────────────────────────
    def _extract_sheet_metadata(self, ws, header_row: int) -> Dict[str, str]:
        """헤더 위 행에서 회사명, 기간 등 추출"""
        metadata = {}

        for r in range(1, header_row):
            for c in range(1, min(10, (ws.max_column or 1) + 1)):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                v = v.strip()

                # 회사명 탐지
                if '주식회사' in v or '(주)' in v:
                    # "주식회사 XXX" 또는 "(주)XXX" 추출
                    m = re.search(r'(?:주식회사\s*|㈜|\(주\))(\S+)', v)
                    if m:
                        metadata['company_name'] = m.group(0).strip()

                # 기간 탐지
                if re.search(r'제?\s*\d+\s*[기(]', v) or re.search(r'\d{4}[년.]\s*\d{1,2}[월.]', v):
                    metadata['fiscal_period'] = v.strip()

                # 단위 탐지
                if '단위' in v:
                    metadata['unit'] = v.strip()

        return metadata

    def _extract_workbook_metadata(self):
        """워크북 수준 메타데이터 통합"""
        for sheet in self.result.sheets:
            if 'company_name' in sheet.metadata and not self.result.company_name:
                self.result.company_name = sheet.metadata['company_name']
            if 'fiscal_period' in sheet.metadata and not self.result.fiscal_period:
                self.result.fiscal_period = sheet.metadata['fiscal_period']

    # ═══════════════════════════════════════════════════════
    # 4. 결과 조회 API
    # ═══════════════════════════════════════════════════════

    def get_trial_balance(self) -> Optional[ParsedSheet]:
        """시산표 시트 반환"""
        for s in self.result.sheets:
            if s.sheet_type == SheetType.TRIAL_BALANCE:
                return s
        return None

    def get_balance_sheet(self) -> Optional[ParsedSheet]:
        """재무상태표 시트 반환"""
        for s in self.result.sheets:
            if s.sheet_type == SheetType.BALANCE_SHEET:
                return s
        return None

    def get_income_statement(self) -> Optional[ParsedSheet]:
        """손익계산서 시트 반환"""
        for s in self.result.sheets:
            if s.sheet_type == SheetType.INCOME_STATEMENT:
                return s
        return None

    def get_all_accounts(self) -> List[Dict[str, Any]]:
        """전체 인식된 계정과목 목록"""
        accounts = []
        for sheet in self.result.sheets:
            for row in sheet.rows:
                if row.recognized_as and not row.is_category:
                    accounts.append({
                        'sheet': sheet.name,
                        'sheet_type': sheet.sheet_type.value,
                        'row': row.row_number,
                        'code': row.account_code,
                        'name': row.account_name,
                        'recognized_as': row.recognized_as,
                        'amounts': row.amounts,
                        'is_subtotal': row.is_subtotal,
                    })
        return accounts

    def get_parse_summary(self) -> Dict[str, Any]:
        """파싱 결과 요약"""
        summary = {
            'filename': self.filepath,
            'company': self.result.company_name,
            'period': self.result.fiscal_period,
            'sheets': [],
        }
        for s in self.result.sheets:
            total = len(s.rows)
            recognized = sum(1 for r in s.rows if r.recognized_as)
            summary['sheets'].append({
                'name': s.name,
                'type': s.sheet_type.value,
                'type_confidence': f"{s.type_confidence:.0%}",
                'header_row': s.header_row,
                'total_rows': total,
                'recognized_accounts': recognized,
                'recognition_rate': f"{recognized/total:.0%}" if total > 0 else "N/A",
                'columns': [
                    {'col': c.letter, 'role': c.role.value, 'header': c.header_text[:20], 'conf': f"{c.confidence:.0%}"}
                    for c in s.columns if c.role != ColumnRole.UNKNOWN
                ],
            })
        return summary


# ═══════════════════════════════════════════════════════════
# 5. CLI 인터페이스
# ═══════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <filepath.xlsx>")
        sys.exit(1)

    parser = ExcelParser(sys.argv[1])
    result = parser.parse()
    summary = parser.get_parse_summary()
    print(json.dumps(summary, ensure_ascii=False, indent=2))
